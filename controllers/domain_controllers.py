"""
controllers/domain_controllers.py
Products / Inventory / Transfer / Logistics / Reports blueprints.
"""
import io
import pandas as pd
from flask import Blueprint, jsonify, request, abort, send_file
from flask_jwt_extended import get_jwt, jwt_required
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from services.services import (
    ProductService, InventoryService,
    TransferService, ShipmentService, ReportService,
)

# ─── Products ────────────────────────────────────────────────────────────────
products_bp = Blueprint('products', __name__)
_prod_svc = ProductService()


@products_bp.route('/', methods=['GET'])
@jwt_required()
def get_products():
    """
    API Lấy danh sách sản phẩm
    ---
    tags:
      - Products
    security:
      - Bearer: []
    responses:
      200:
        description: Danh sách sản phẩm
    """
    return jsonify(_prod_svc.get_all()), 200


@products_bp.route('/<id>', methods=['GET'])
@jwt_required()
def get_product_id(id):
    """
    API Lấy sản phẩm theo ID
    ---
    tags:
      - Products
    security:
      - Bearer: []
    parameters:
      - name: id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: Thông tin sản phẩm
    """
    return jsonify(_prod_svc.get_by_id(id)), 200


@products_bp.route('/name=<name>', methods=['GET'])
@jwt_required()
def get_product_name(name):
    """
    API Tìm sản phẩm theo tên
    ---
    tags:
      - Products
    security:
      - Bearer: []
    parameters:
      - name: name
        in: path
        required: true
        type: string
    responses:
      200:
        description: Danh sách sản phẩm khớp tên
    """
    return jsonify(_prod_svc.get_by_name(name)), 200


@products_bp.route('/', methods=['POST'])
@jwt_required()
def add_product():
    """
    API Thêm sản phẩm mới
    ---
    tags:
      - Products
    security:
      - Bearer: []
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            sku:
              type: string
            product_name:
              type: string
            category_name:
              type: string
            min_stock:
              type: integer
            description:
              type: string
    responses:
      200:
        description: Thêm thành công
    """
    _prod_svc.create(request.get_json() or {})
    return jsonify({'success': True, 'message': 'Sản phẩm đã được thêm vào'}), 200


@products_bp.route('/<id>', methods=['PUT'])
@jwt_required()
def update_product(id):
    """
    API Cập nhật sản phẩm
    ---
    tags:
      - Products
    security:
      - Bearer: []
    parameters:
      - name: id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: Cập nhật thành công
      403:
        description: Không có quyền
    """
    if get_jwt().get("role") not in ("MANAGER", "ADMIN"):
        abort(403, description="Bạn không có quyền thực hiện hành động này")
    _prod_svc.update(id, request.get_json() or {})
    return jsonify({'success': True, 'message': 'Thông tin sản phẩm đã được cập nhật'}), 200


@products_bp.route('/<id>', methods=['DELETE'])
@jwt_required()
def delete_product(id):
    """
    API Xóa sản phẩm
    ---
    tags:
      - Products
    security:
      - Bearer: []
    parameters:
      - name: id
        in: path
        required: true
        type: integer
    responses:
      200:
        description: Đã xóa
      403:
        description: Không có quyền
    """
    if get_jwt().get("role") not in ("MANAGER", "ADMIN"):
        abort(403, description="Bạn không có quyền thực hiện hành động này")
    _prod_svc.delete(id)
    return jsonify({'success': True, 'message': 'Thông tin sản phẩm đã được xóa'}), 200


# ─── Inventory ───────────────────────────────────────────────────────────────
inventory_bp = Blueprint('inventory', __name__)
_inv_svc = InventoryService()


@inventory_bp.route('/', methods=['GET'])
def get_inventory():
    return jsonify({"success": True, "data": _inv_svc.get_all()})


@inventory_bp.route('/add', methods=['POST'])
def add_inventory():
    data = request.json or {}
    _inv_svc.add(data.get("warehouse_name"), data.get("product_name"), data.get("quantity"))
    return jsonify({"success": True, "message": "Added successfully"})


@inventory_bp.route('/remove', methods=['POST'])
def remove_inventory():
    data = request.json or {}
    _inv_svc.remove(data.get("warehouse_name"), data.get("product_name"), data.get("quantity"))
    return jsonify({"success": True, "message": "Removed successfully"})


@inventory_bp.route('/delete/<int:id>', methods=['DELETE'])
def delete_inventory(id):
    _inv_svc.delete(id)
    return jsonify({"success": True, "message": "Deleted successfully"})


@inventory_bp.route('/update/<int:id>', methods=['PUT'])
def update_inventory(id):
    data = request.json or {}
    if "quantity" not in data:
        abort(400, "Missing quantity")
    _inv_svc.update(id, data['quantity'])
    return jsonify({"success": True, "message": "Updated successfully"})


@inventory_bp.route('/search/product', methods=['GET'])
def search_product():
    name = request.args.get("name")
    return jsonify({"success": True, "data": _inv_svc.search_by_product(name)})


@inventory_bp.route('/search/warehouse', methods=['GET'])
def search_warehouse():
    name = request.args.get("name")
    return jsonify({"success": True, "data": _inv_svc.search_by_warehouse(name)})


# ─── Transfers ───────────────────────────────────────────────────────────────
transfer_bp = Blueprint('transfer', __name__)
_transfer_svc = TransferService()


@transfer_bp.route('/', methods=['GET'])
def get_transfers():
    """
    Lấy tất cả Transfer Orders
    ---
    tags:
      - Transfers
    responses:
      200:
        description: Success
    """
    return jsonify({"success": True, "data": _transfer_svc.get_all()})


@transfer_bp.route('/transfers', methods=['POST'])
def create_transfer():
    """
    Tạo Transfer Order
    ---
    tags:
      - Transfers
    parameters:
      - in: body
        name: body
        schema:
          type: object
          required:
            - from_warehouse_id
            - to_warehouse_id
            - staff_id
            - products
          properties:
            from_warehouse_id:
              type: integer
            to_warehouse_id:
              type: integer
            staff_id:
              type: integer
            products:
              type: array
              items:
                type: object
    responses:
      200:
        description: Created
    """
    data = request.json or {}
    required = ["from_warehouse_id", "to_warehouse_id", "staff_id", "products"]
    if not all(k in data for k in required):
        abort(400, "Missing fields")
    transfer_id = _transfer_svc.create(
        data['from_warehouse_id'], data['to_warehouse_id'],
        data['staff_id'], data['products']
    )
    return jsonify({"success": True, "data": {"transfer_id": transfer_id, "status": "PENDING"}})


@transfer_bp.route('/transfers/<int:id>', methods=['PUT'])
def update_transfer_status(id):
    """
    Cập nhật trạng thái Transfer
    ---
    tags:
      - Transfers
    parameters:
      - in: path
        name: id
        type: integer
        required: true
      - in: body
        name: body
        schema:
          type: object
          properties:
            status:
              type: string
    responses:
      200:
        description: Updated
    """
    data = request.json or {}
    if 'status' not in data:
        abort(400, "Missing status")
    _transfer_svc.update_status(id, data['status'])
    return jsonify({"success": True, "data": {"transfer_id": id, "status": data['status']}})


@transfer_bp.route('/transfers/suggest', methods=['GET'])
def suggest_warehouse():
    """
    Gợi ý kho tốt nhất
    ---
    tags:
      - Transfers
    parameters:
      - in: query
        name: product_id
        type: integer
        required: true
      - in: query
        name: to_warehouse_id
        type: integer
        required: true
    responses:
      200:
        description: Success
    """
    product_id = request.args.get('product_id', type=int)
    to_warehouse_id = request.args.get('to_warehouse_id', type=int)
    if not product_id or not to_warehouse_id:
        abort(400, "Missing params")
    best = _transfer_svc.suggest_warehouse(product_id, to_warehouse_id)
    return jsonify({"success": True, "data": best})


# ─── Shipments (Logistics) ───────────────────────────────────────────────────
logistics_bp = Blueprint('logistics', __name__)
_ship_svc = ShipmentService()


@logistics_bp.route('/shipments', methods=['GET'])
def get_shipments():
    """
    Lấy tất cả Shipments
    ---
    tags:
      - Shipments
    responses:
      200:
        description: Danh sách shipments
    """
    return jsonify({"success": True, "data": _ship_svc.get_all()})


@logistics_bp.route('/shipments', methods=['POST'])
def create_shipment():
    """
    Tạo Shipment
    ---
    tags:
      - Shipments
    parameters:
      - in: body
        name: body
        schema:
          type: object
          required:
            - transfer_id
            - driver_name
            - license_plate
            - expected_delivery_at
    responses:
      200:
        description: Created
    """
    data = request.json or {}
    required = ["transfer_id", "driver_name", "license_plate", "expected_delivery_at"]
    if not all(k in data for k in required):
        abort(400, "Missing fields")
    _ship_svc.create(
        data['transfer_id'], data['driver_name'],
        data['license_plate'], data['expected_delivery_at']
    )
    return jsonify({"success": True, "data": "Shipment created"})


@logistics_bp.route('/shipments/<int:id>', methods=['PUT'])
def update_shipment(id):
    """
    Cập nhật Shipment
    ---
    tags:
      - Shipments
    parameters:
      - in: path
        name: id
        type: integer
        required: true
    responses:
      200:
        description: Updated
    """
    data = request.json or {}
    if not data:
        abort(400, "Invalid JSON")
    _ship_svc.update(id, data)
    return jsonify({"success": True, "data": "Shipment updated"})


@logistics_bp.route('/shipments/<int:id>', methods=['DELETE'])
def delete_shipment(id):
    """
    Xóa Shipment
    ---
    tags:
      - Shipments
    parameters:
      - in: path
        name: id
        type: integer
        required: true
    responses:
      200:
        description: Deleted
    """
    _ship_svc.delete(id)
    return jsonify({"success": True, "data": "Shipment deleted"})


@logistics_bp.route('/shipments/<int:id>', methods=['GET'])
def get_shipment_by_id(id):
    """
    Lấy Shipment theo ID
    ---
    tags:
      - Shipments
    parameters:
      - in: path
        name: id
        type: integer
        required: true
    responses:
      200:
        description: Thông tin shipment
    """
    return jsonify({"success": True, "data": _ship_svc.get_by_id(id)})


# ─── Reports ─────────────────────────────────────────────────────────────────
reports_bp = Blueprint('reports', __name__)
_report_svc = ReportService()


def _success(data):
    import flask
    return flask.jsonify({
        "success": True,
        "count": len(data) if isinstance(data, list) else 1,
        "data": data
    }), 200


@reports_bp.route('', methods=['GET'])
@jwt_required()
def get_ton_kho():
    """
    API lấy tồn kho
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    responses:
      200:
        description: Tồn kho
    """
    return _success(_report_svc.get_ton_kho())


@reports_bp.route('/low-stock', methods=['GET'])
@jwt_required()
def get_low_stock():
    """
    API sản phẩm sắp hết
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    responses:
      200:
        description: Sản phẩm sắp hết
    """
    return _success(_report_svc.get_low_stock())


@reports_bp.route('/history/<string:sku>', methods=['GET'])
@jwt_required()
def get_inventory_history(sku):
    """
    API lịch sử nhập/xuất theo SKU
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    parameters:
      - name: sku
        in: path
        required: true
        type: string
    responses:
      200:
        description: Lịch sử
    """
    return _success(_report_svc.get_inventory_history(sku))


@reports_bp.route('/transfer-history', methods=['GET'])
@jwt_required()
def get_transfer_history():
    """
    API lịch sử điều chuyển
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    responses:
      200:
        description: Lịch sử điều chuyển
    """
    import flask
    return _success(_report_svc.get_transfer_history(flask.request.args.get('sku')))


@reports_bp.route('/receipt', methods=['GET'])
@jwt_required()
def get_receipt_history():
    """
    API lịch sử phiếu nhập/xuất
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    responses:
      200:
        description: Lịch sử phiếu
    """
    return _success(_report_svc.get_receipt_history())


@reports_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    """
    API export Excel
    ---
    tags:
      - Reports
    security:
      - Bearer: []
    produces:
      - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    responses:
      200:
        description: File Excel
    """
    excel_data = _report_svc.get_excel_data()
    receipts = excel_data["receipts"]

    df5 = pd.DataFrame(receipts)
    if not df5.empty:
        total_row = {col: '' for col in df5.columns}
        total_row['product_name'] = 'TOTAL'
        total_row['total_value'] = df5['total_value'].sum()
        df5 = pd.concat([df5, pd.DataFrame([total_row])], ignore_index=True)

    sheets = {
        'TonKho': pd.DataFrame(excel_data["inventory"]),
        'LowStock': pd.DataFrame(excel_data["stock_summary"]),
        'History': pd.DataFrame(excel_data["history"]),
        'Transfer': pd.DataFrame(excel_data["transfers"]),
        'Receipts': df5,
    }

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            for col_idx, col in enumerate(df.columns, 1):
                max_len = max(
                    len(str(col)),
                    *[len(str(c)[:19] if "created_at" in col.lower() else str(c))
                      for c in df[col] if c]
                ) if not df.empty else len(str(col))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
    output.seek(0)
    return send_file(
        output,
        download_name="inventory_report.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )